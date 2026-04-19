from django.shortcuts import render, redirect
from django.http import HttpResponse
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import login_required
from django.contrib.auth.models import User
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, TableStyle, Table
from reportlab.lib.styles import getSampleStyleSheet
from .models import UserProfile, Activity
from reportlab.lib import colors
from reportlab.lib.styles import ParagraphStyle
from reportlab.lib.pagesizes import letter
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from .forms import UserUpdateForm, ProfileUpdateForm
import os
from django.conf import settings

import joblib
import pandas as pd
import json

# Load model
model = joblib.load("model.pkl")



def login_view(request):
    if request.method == "POST":
        user = authenticate(
            request,
            username=request.POST["username"],
            password=request.POST["password"]
        )
        if user:
            login(request, user)
            log_activity(user, 'login', f"User logged in from IP: {request.META.get('REMOTE_ADD R')}")

            try: 
                profile = user.userprofile
                if profile.role == 'admin':
                    return redirect("/dashboard/")
                else:
                    return redirect("/")
            except UserProfile.DoesNotExist:
                UserProfile.objects.create(user=user, role='normal')
                return redirect("/")        
        
        else:
          return render(request, "login.html", {"error": "Invalid credentials"})
    return render(request, "login.html")


def logout_view(request):
    if request.user.is_authenticated:
        log_activity(request.user, 'logout', "User logged out")
    logout(request)
    return redirect("/login/")


@login_required
def home(request):
    is_admin = request.user.userprofile.role == 'admin' if hasattr(request.user, 'userprofile') else False
   
    # This is for FILTERING existing results (after prediction)
    if request.method == "GET" and request.GET.get("risk"):
        df = pd.read_excel("results.xlsx")
        
        risk = request.GET.get("risk")
        if risk:
            df = df[df["Risk"] == risk]
        
        min_age = request.GET.get("min_age")
        max_age = request.GET.get("max_age")
        
        if min_age:
            df = df[df["age"] >= int(min_age)]
        if max_age:
            df = df[df["age"] <= int(max_age)]
        
        high_count = (df["Risk"] == "High").sum()
        low_count = (df["Risk"] == "Low").sum()
        
        chart_data = {
            "high": int(high_count),
            "low": int(low_count)
        }
        
        return render(request, "result.html", {
            "table": df.to_html(classes="table table-striped table-bordered"),
            "chart_data": json.dumps(chart_data),
            "total_patients": len(df),
            "high_risk": high_count,
            "low_risk": low_count
        })
    
    # This is for UPLOAD and PREDICTION
    if request.method == "POST" and request.FILES.get("file"):
        file = request.FILES["file"]
        
        # Read file
        if file.name.endswith('.csv'):
            df = pd.read_csv(file)
        elif file.name.endswith('.xlsx'):
            df = pd.read_excel(file)
        else:
            return HttpResponse("Unsupported file format")
        
        # Log the upload
        log_activity(request.user, 'upload', f"Uploaded file: {file.name} with {len(df)} patients")
        
        # Preprocessing - fill missing values
        df.fillna(df.mean(numeric_only=True), inplace=True)
        
        expected_columns = [
            "age","sex","cp","trestbps","chol","fbs",
            "restecg","thalach","exang","oldpeak",
            "slope","ca","thal"
        ]
        
        # Ensure all columns exist
        for col in expected_columns:
            if col not in df.columns:
                df[col] = 0
        
        # Keep original data for reference
        original_df = df.copy()
        df = df[expected_columns]
        
        # Make PREDICTIONS 
        predictions = model.predict(df)
        probabilities = model.predict_proba(df)[:, 1] * 100
        
        # ADD predictions to the dataframe
        original_df["Risk"] = ["High" if p == 1 else "Low" for p in predictions]
        original_df["Risk_Probability_%"] = probabilities.round(2)
        
        # Save results with predictions
        original_df.to_excel("results.xlsx", index=False)
        
        # Calculate statistics
        high_count = sum(predictions)
        low_count = len(predictions) - high_count
        
        chart_data = {
            "high": int(high_count),
            "low": int(low_count)
        }
        high_count = sum(predictions)
        low_count = len(predictions) - high_count
        high_risk_percentage = (high_count / len(predictions)) * 100 if len(predictions) > 0 else 0
        low_risk_percentage = (low_count / len(predictions)) * 100 if len(predictions) > 0 else 0
        risk_ratio = f"{high_count}:{low_count}" if low_count > 0 else f"{high_count}:0"
        
        # Log the prediction
        log_activity(request.user, 'predict', f"Made predictions on {len(df)} patients - {high_count} high risk, {low_count} low risk")
        
        
        return render(request, "result.html", {
               "table": original_df.to_html(classes="table table-striped table-bordered"),
               "chart_data": json.dumps(chart_data),
               "total_patients": len(original_df),
               "high_risk": high_count,
               "low_risk": low_count,
               "high_risk_percentage": round(high_risk_percentage, 1),
               "low_risk_percentage": round(low_risk_percentage, 1),
               "risk_ratio": risk_ratio,
               "filename": file.name
        })
    
    return render(request, "home.html", {'is_admin': is_admin})


def signup(request):
    if request.method == "POST":
        username = request.POST["username"]
        email = request.POST["email"]
        password = request.POST["password"]
        role = request.POST.get("role", "normal")

        print(f"Signup role: {role}")

        # prevent duplicate users
        if User.objects.filter(username=username).exists():
            return render(request, "signup.html", {"error": "Username already exists"})
        if User.objects.filter(email=email).exists():
            return render(request, "signup.html", {"error: Email already created"})

        
        user = User.objects.create_user(username=username, password=password, email=email)
        UserProfile.objects.create(user=user, role=role)

        login(request, user)
    
    
        if role == 'admin':
            return redirect("/dashboard/")
        else:
            return redirect("/")

    return render(request, "signup.html")

def is_admin(user):
    """Helper function to check if user is admin"""
    return hasattr(user, 'userprofile') and user.userprofile.role == 'admin'

@login_required
def download(request):
    try:
        df = pd.read_excel("results.xlsx")
    except:
        return HttpResponse("No data available. Please upload and analyze data first.")
    
    # Create a new Excel writer with multiple sheets
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="clinical_risk_report.xlsx"'
    
    # Use pandas ExcelWriter to create multiple sheets
    with pd.ExcelWriter(response, engine='openpyxl') as writer:
        
        # Sheet 1: Executive Summary
        summary_data = {
            'Metric': ['Report Generated', 'Total Patients Analyzed', 'High Risk Patients', 'Low Risk Patients', 
                      'Risk Ratio (High:Low)', 'High Risk Percentage', 'Low Risk Percentage'],
            'Value': [
                pd.Timestamp.now().strftime('%Y-%m-%d %H:%M:%S'),
                len(df),
                (df["Risk"] == "High").sum(),
                (df["Risk"] == "Low").sum(),
                f"{(df['Risk'] == 'High').sum()}:{(df['Risk'] == 'Low').sum()}",
                f"{((df['Risk'] == 'High').sum() / len(df) * 100):.1f}%",
                f"{((df['Risk'] == 'Low').sum() / len(df) * 100):.1f}%"
            ]
        }
        summary_df = pd.DataFrame(summary_data)
        summary_df.to_excel(writer, sheet_name='Executive Summary', index=False)
        
        # Sheet 2: Clinical Insights
        high_count = (df["Risk"] == "High").sum()
        low_count = (df["Risk"] == "Low").sum()
        total = len(df)
        
        if high_count > total / 2:
            risk_level = "CRITICAL RISK LEVEL"
            finding = "More than 50% of patients are at HIGH RISK for cardiovascular disease"
        elif high_count > total / 3:
            risk_level = "HIGH RISK LEVEL"
            finding = f"Significant portion ({high_count} out of {total}) show elevated risk markers"
        elif high_count > 0:
            risk_level = "MODERATE RISK LEVEL"
            finding = f"{high_count} patient(s) identified as high risk"
        else:
            risk_level = "LOW RISK LEVEL"
            finding = "No high-risk patients detected"
        
        insights_data = {
            'Category': ['Overall Risk Classification', 'Key Finding', 'Clinical Recommendation'],
            'Details': [
                risk_level,
                finding,
                'Immediate cardiology referral for high-risk patients' if high_count > 10 else 
                'Follow-up assessment within 1-2 months' if high_count > 0 else 
                'Continue preventive health measures'
            ]
        }
        insights_df = pd.DataFrame(insights_data)
        insights_df.to_excel(writer, sheet_name='Clinical Insights', index=False)
        
        # Sheet 3: Risk Factor Analysis
        high_risk_patients = df[df["Risk"] == "High"]
        low_risk_patients = df[df["Risk"] == "Low"]
        
        risk_analysis = {
            'Risk Factor': ['Average Age', 'Average Cholesterol', 'Average Max Heart Rate', 'Average ST Depression'],
            'High-Risk Patients': [
                round(high_risk_patients['age'].mean(), 1) if len(high_risk_patients) > 0 else 0,
                round(high_risk_patients['chol'].mean(), 1) if len(high_risk_patients) > 0 else 0,
                round(high_risk_patients['thalach'].mean(), 1) if len(high_risk_patients) > 0 else 0,
                round(high_risk_patients['oldpeak'].mean(), 1) if len(high_risk_patients) > 0 else 0
            ],
            'Low-Risk Patients': [
                round(low_risk_patients['age'].mean(), 1) if len(low_risk_patients) > 0 else 0,
                round(low_risk_patients['chol'].mean(), 1) if len(low_risk_patients) > 0 else 0,
                round(low_risk_patients['thalach'].mean(), 1) if len(low_risk_patients) > 0 else 0,
                round(low_risk_patients['oldpeak'].mean(), 1) if len(low_risk_patients) > 0 else 0
            ]
        }
        risk_df = pd.DataFrame(risk_analysis)
        risk_df.to_excel(writer, sheet_name='Risk Factor Analysis', index=False)
        
        # Sheet 4: Detailed Predictions (Original data with risk assessment)
        df.to_excel(writer, sheet_name='Patient Predictions', index=False)
        
        # Sheet 5: High-Risk Patients List (Filtered)
        if len(high_risk_patients) > 0:
            high_risk_df = high_risk_patients[['age', 'sex', 'chol', 'thalach', 'oldpeak', 'Risk', 'Risk_Probability_%']]
            high_risk_df.to_excel(writer, sheet_name='High Risk Patients', index=False)
        
        # Sheet 6: Recommendations
        recommendations = {
            'Priority': ['Immediate', 'Short-term', 'Long-term', 'Monitoring'],
            'Action Items': [
                'Schedule cardiology consultation for all high-risk patients',
                'Conduct additional diagnostic tests (Echocardiogram, Stress Test)',
                'Implement lifestyle intervention program (diet, exercise, smoking cessation)',
                'Establish 3-month follow-up protocol for high-risk patients'
            ],
            'Responsible Party': ['Primary Care Physician', 'Cardiology Department', 'Patient + Health Coach', 'Clinic Staff']
        }
        recommendations_df = pd.DataFrame(recommendations)
        recommendations_df.to_excel(writer, sheet_name='Recommendations', index=False)
        
        # Add charts using openpyxl
        from openpyxl import load_workbook
        from openpyxl.chart import PieChart, BarChart, Reference
        from openpyxl.chart.label import DataLabelList
        
        # Get the workbook
        workbook = writer.book
        
        # Create Pie Chart for Risk Distribution
        chart_sheet = workbook.create_sheet("Risk Chart")
        
        # Add data for chart
        chart_sheet['A1'] = 'Risk Category'
        chart_sheet['B1'] = 'Count'
        chart_sheet['A2'] = 'High Risk'
        chart_sheet['B2'] = high_count
        chart_sheet['A3'] = 'Low Risk'
        chart_sheet['B3'] = low_count
        
        # Create pie chart
        pie_chart = PieChart()
        data = Reference(chart_sheet, min_col=2, min_row=1, max_row=3, max_col=2)
        labels = Reference(chart_sheet, min_col=1, min_row=2, max_row=3)
        pie_chart.add_data(data, titles_from_data=True)
        pie_chart.set_categories(labels)
        pie_chart.title = "Risk Distribution"
        pie_chart.width = 15
        pie_chart.height = 10
        
        # Add data labels
        pie_chart.dataLabels = DataLabelList()
        pie_chart.dataLabels.showVal = True
        pie_chart.dataLabels.showPercent = True
        
        chart_sheet.add_chart(pie_chart, "E2")
        
        # Create Bar Chart for Risk Factor Comparison
        bar_chart = BarChart()
        bar_chart.title = "Risk Factor Comparison"
        bar_chart.style = 10
        bar_chart.width = 20
        bar_chart.height = 12
        
        # Add data for bar chart
        bar_sheet = workbook.create_sheet("Risk Factor Chart")
        bar_sheet['A1'] = 'Risk Factor'
        bar_sheet['B1'] = 'High Risk'
        bar_sheet['C1'] = 'Low Risk'
        
        risk_factors = ['Average Age', 'Avg Cholesterol', 'Avg Max HR', 'Avg ST Depression']
        high_values = [
            round(high_risk_patients['age'].mean(), 1) if len(high_risk_patients) > 0 else 0,
            round(high_risk_patients['chol'].mean(), 1) if len(high_risk_patients) > 0 else 0,
            round(high_risk_patients['thalach'].mean(), 1) if len(high_risk_patients) > 0 else 0,
            round(high_risk_patients['oldpeak'].mean(), 1) if len(high_risk_patients) > 0 else 0
        ]
        low_values = [
            round(low_risk_patients['age'].mean(), 1) if len(low_risk_patients) > 0 else 0,
            round(low_risk_patients['chol'].mean(), 1) if len(low_risk_patients) > 0 else 0,
            round(low_risk_patients['thalach'].mean(), 1) if len(low_risk_patients) > 0 else 0,
            round(low_risk_patients['oldpeak'].mean(), 1) if len(low_risk_patients) > 0 else 0
        ]
        
        for i, factor in enumerate(risk_factors, start=2):
            bar_sheet[f'A{i}'] = factor
            bar_sheet[f'B{i}'] = high_values[i-2]
            bar_sheet[f'C{i}'] = low_values[i-2]
        
        data = Reference(bar_sheet, min_col=2, min_row=1, max_row=len(risk_factors)+1, max_col=3)
        categories = Reference(bar_sheet, min_col=1, min_row=2, max_row=len(risk_factors)+1)
        bar_chart.add_data(data, titles_from_data=True)
        bar_chart.set_categories(categories)
        
        bar_sheet.add_chart(bar_chart, "E2")
        
        # Log the download activity
        log_activity(request.user, 'download', f"Downloaded comprehensive Excel report with {total} patients analyzed")
    
    return response


@login_required
def download_pdf(request):
    try:
        df = pd.read_excel("results.xlsx")
    except:
        return HttpResponse("No data available. Please upload and analyze data first.")
    
    # Calculate statistics
    total = len(df)
    high = (df["Risk"] == "High").sum()
    low = (df["Risk"] == "Low").sum()
    high_percentage = (high / total * 100) if total > 0 else 0
    low_percentage = (low / total * 100) if total > 0 else 0
    
    # Calculate average risk factors for high-risk patients
    high_risk_patients = df[df["Risk"] == "High"]
    low_risk_patients = df[df["Risk"] == "Low"]
    
    avg_age_high = round(high_risk_patients['age'].mean(), 1) if len(high_risk_patients) > 0 else 0
    avg_age_low = round(low_risk_patients['age'].mean(), 1) if len(low_risk_patients) > 0 else 0
    avg_chol_high = round(high_risk_patients['chol'].mean(), 1) if len(high_risk_patients) > 0 else 0
    avg_chol_low = round(low_risk_patients['chol'].mean(), 1) if len(low_risk_patients) > 0 else 0
    
    # Create PDF response
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="clinical_risk_assessment_report.pdf"'
    
    # Create PDF document
    doc = SimpleDocTemplate(response, pagesize=(612, 936))  # Letter size
    styles = getSampleStyleSheet()
    
    # Create custom styles
    styles.add(ParagraphStyle(name='CustomTitle', 
                              parent=styles['Title'],
                              fontSize=24, 
                              textColor=colors.HexColor('#667eea'),
                              alignment=1,  # Center alignment
                              spaceAfter=30))
    
    styles.add(ParagraphStyle(name='SectionHeader',
                              parent=styles['Heading2'],
                              fontSize=16,
                              textColor=colors.HexColor('#374151'),
                              spaceBefore=20,
                              spaceAfter=10))
    
    styles.add(ParagraphStyle(name='RiskHigh',
                              parent=styles['Normal'],
                              textColor=colors.HexColor('#dc2626'),
                              fontSize=12))
    
    styles.add(ParagraphStyle(name='RiskLow',
                              parent=styles['Normal'],
                              textColor=colors.HexColor('#10b981'),
                              fontSize=12))
    
    styles.add(ParagraphStyle(name='SummaryText',
                              parent=styles['Normal'],
                              fontSize=11,
                              leading=16))
    
    elements = []
    
    # Title
    elements.append(Paragraph("Cardiovascular Risk Assessment Report", styles['CustomTitle']))
    elements.append(Spacer(1, 10))
    
    # Date
    from datetime import datetime
    elements.append(Paragraph(f"Generated: {datetime.now().strftime('%B %d, %Y at %I:%M %p')}", 
                              styles['Normal']))
    elements.append(Spacer(1, 20))
    
    # EXECUTIVE SUMMARY
    elements.append(Paragraph("Executive Summary", styles['SectionHeader']))
    
    if high > total / 2:
        summary = "⚠️ CRITICAL FINDING: More than 50% of patients in this cohort are at HIGH RISK for cardiovascular disease. Immediate medical attention and lifestyle interventions are strongly recommended."
    elif high > total / 3:
        summary = f"⚠️ MODERATE-HIGH ALERT: A significant portion ({high} out of {total}) of patients show elevated risk markers. Preventive cardiology consultation advised for these individuals."
    elif high > 0:
        summary = f"📊 OBSERVATION: {high} patient(s) identified as high risk. Targeted interventions and regular monitoring recommended."
    else:
        summary = "✅ FAVORABLE OUTCOME: No high-risk patients detected in this analysis. Continue preventive health measures."
    
    elements.append(Paragraph(summary, styles['SummaryText']))
    elements.append(Spacer(1, 15))
    
    # RISK LEVEL
    if high_percentage > 50:
        risk_level = "CRITICAL RISK LEVEL"
        risk_color = colors.HexColor('#dc2626')
    elif high_percentage > 30:
        risk_level = "HIGH RISK LEVEL"
        risk_color = colors.HexColor('#f59e0b')
    elif high_percentage > 10:
        risk_level = "MODERATE RISK LEVEL"
        risk_color = colors.HexColor('#3b82f6')
    else:
        risk_level = "LOW RISK LEVEL"
        risk_color = colors.HexColor('#10b981')
    
    elements.append(Paragraph(f"Overall Risk Classification: {risk_level}", 
                              ParagraphStyle(name='RiskLevel', parent=styles['Normal'],
                                           fontSize=14, textColor=risk_color, fontName='Helvetica-Bold')))
    elements.append(Spacer(1, 20))
    
    # KEY STATISTICS
    elements.append(Paragraph("Key Statistics", styles['SectionHeader']))
    
    stats_data = [
        ["Total Patients Analyzed", str(total)],
        ["High Risk Patients", f"{high} ({high_percentage:.1f}%)"],
        ["Low Risk Patients", f"{low} ({low_percentage:.1f}%)"],
        ["Risk Ratio (High:Low)", f"{high}:{low}" if low > 0 else f"{high}:0"],
    ]
    
    stats_table = Table(stats_data, colWidths=[250, 150])
    stats_table.setStyle(TableStyle([
        ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 0), (-1, -1), 11),
        ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#f3f4f6')),
        ('TEXTCOLOR', (0, 0), (0, -1), colors.HexColor('#374151')),
        ('ALIGN', (1, 0), (1, -1), 'CENTER'),
        ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#e5e7eb')),
        ('TOPPADDING', (0, 0), (-1, -1), 8),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
    ]))
    elements.append(stats_table)
    elements.append(Spacer(1, 20))
    
    # CLINICAL INSIGHTS
    elements.append(Paragraph("Clinical Insights & Recommendations", styles['SectionHeader']))
    
    insights = []
    if high > 0:
        insights.append(f"• High-Risk Patient Profile: {high} patients ({high_percentage:.1f}%) exhibit significant risk markers for cardiovascular events. Key indicators include elevated cholesterol, hypertension markers, and abnormal ECG findings in this subgroup.")
    
    insights.append(f"• Population Risk Distribution: The risk profile shows {'predominantly high-risk cases' if high > low else 'predominantly low-risk cases'}, suggesting {'an urgent need for population-level interventions' if high > low else 'effective current health management'}.")
    
    if high > 10:
        insights.append("• Clinical Recommendation: High-risk patients require immediate cardiology referral, lifestyle modification program, and regular monitoring (every 3-6 months).")
    elif high > 0:
        insights.append("• Clinical Recommendation: High-risk patients should be scheduled for follow-up assessments within 1-2 months for preventive intervention.")
    else:
        insights.append("• Clinical Recommendation: Continue current preventive health strategies and maintain annual screening for early detection.")
    
    for insight in insights:
        elements.append(Paragraph(insight, styles['SummaryText']))
        elements.append(Spacer(1, 8))
    
    elements.append(Spacer(1, 15))
    
    # RISK FACTOR ANALYSIS
    elements.append(Paragraph("Key Risk Factor Analysis", styles['SectionHeader']))
    
    risk_factors_data = [
        ["Risk Factor", "High-Risk Patients", "Low-Risk Patients"],
        ["Average Age", f"{avg_age_high} years", f"{avg_age_low} years"],
        ["Average Cholesterol", f"{avg_chol_high} mg/dL", f"{avg_chol_low} mg/dL"],
    ]
    
    if len(high_risk_patients) > 0 and 'thalach' in high_risk_patients.columns:
        avg_hr_high = round(high_risk_patients['thalach'].mean(), 1)
        avg_hr_low = round(low_risk_patients['thalach'].mean(), 1) if len(low_risk_patients) > 0 else 0
        risk_factors_data.append(["Avg Max Heart Rate", f"{avg_hr_high} bpm", f"{avg_hr_low} bpm"])
    
    if len(high_risk_patients) > 0 and 'oldpeak' in high_risk_patients.columns:
        avg_st_high = round(high_risk_patients['oldpeak'].mean(), 1)
        avg_st_low = round(low_risk_patients['oldpeak'].mean(), 1) if len(low_risk_patients) > 0 else 0
        risk_factors_data.append(["ST Depression (oldpeak)", f"{avg_st_high}", f"{avg_st_low}"])
    
    risk_table = Table(risk_factors_data, colWidths=[150, 150, 150])
    risk_table.setStyle(TableStyle([
        ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#667eea')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('ALIGN', (1, 0), (-1, -1), 'CENTER'),
        ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#e5e7eb')),
        ('TOPPADDING', (0, 0), (-1, -1), 8),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
    ]))
    elements.append(risk_table)
    elements.append(Spacer(1, 20))
    
    # RECOMMENDATIONS
    elements.append(Paragraph("Immediate Action Items", styles['SectionHeader']))
    
    recommendations = [
        "• Schedule cardiology consultation for all high-risk patients",
        "• Conduct additional diagnostic tests (Echocardiogram, Stress Test) for high-risk cases",
        "• Review and optimize medication regimens for hypertension and cholesterol management",
        "• Implement lifestyle intervention program (diet, exercise, smoking cessation)",
        "• Establish 3-month follow-up protocol for high-risk patient monitoring"
    ]
    
    for rec in recommendations:
        elements.append(Paragraph(rec, styles['SummaryText']))
        elements.append(Spacer(1, 6))
    
    elements.append(Spacer(1, 20))
    
    # HIGH-RISK PATIENT LIST 
    if len(high_risk_patients) > 0:
        elements.append(Paragraph("High-Risk Patients Needing Immediate Attention", styles['SectionHeader']))
        
        high_risk_list = []
        for idx, (_, patient) in enumerate(high_risk_patients.head(10).iterrows(), 1):
            age = patient.get('age', 'N/A')
            chol = patient.get('chol', 'N/A')
            prob = patient.get('Risk_Probability_%', 'N/A')
            high_risk_list.append(f"{idx}. Age: {age}, Cholesterol: {chol}, Risk Probability: {prob}%")
        
        if len(high_risk_patients) > 10:
            high_risk_list.append(f"... and {len(high_risk_patients) - 10} more high-risk patients")
        
        for patient_info in high_risk_list:
            elements.append(Paragraph(patient_info, styles['RiskHigh']))
            elements.append(Spacer(1, 4))
        
        elements.append(Spacer(1, 15))
    

    elements.append(Paragraph("Disclaimer", ParagraphStyle(name='Disclaimer', parent=styles['Normal'],
                                                          fontSize=8, textColor=colors.HexColor('#9ca3af'),
                                                          alignment=1)))
    elements.append(Paragraph("This report is generated by an AI-based risk assessment tool and should not replace professional medical advice. "
                              "All clinical decisions should be made in consultation with qualified healthcare providers.", 
                              ParagraphStyle(name='DisclaimerText', parent=styles['Normal'],
                                           fontSize=8, textColor=colors.HexColor('#9ca3af'), alignment=1)))
    
    
    doc.build(elements)
    return response

@login_required
def dashboard(request):
    if is_admin(request.user):
        
        log_activity(request.user, 'view_dashboard', "Viewed admin dashboard")
        
        try:
            df = pd.read_excel("results.xlsx")
            total = len(df)
            high = (df["Risk"] == "High").sum()
            low = (df["Risk"] == "Low").sum()
        except:
            total, high, low = 0, 0, 0
        
     
        recent_activities = Activity.objects.all()[:50]
        
     
        from django.db.models import Count
        activity_stats = Activity.objects.values('action').annotate(count=Count('action'))
        
      
        user_activity_stats = Activity.objects.values('user__username', 'user__userprofile__role')\
            .annotate(activity_count=Count('id'))\
            .order_by('-activity_count')[:10]
        
        return render(request, "dashboard.html", {
            "total": total,
            "high": high,
            "low": low,
            "recent_activities": recent_activities,
            "activity_stats": activity_stats,
            "user_activity_stats": user_activity_stats,
        })
    else:
        return render(request, "access_denied.html")
    
def log_activity(user, action, details=None):
    Activity.objects.create(
        user=user,
        action=action,
        details=details
    )

@login_required
def profile_view(request):
    """View user profile"""
    user_profile, created = UserProfile.objects.get_or_create(user=request.user)
    
    context = {
        'user_profile': user_profile,
    }
    return render(request, 'profile.html', context)

@login_required
def profile_edit(request):
    """Edit user profile"""
    user_profile, created = UserProfile.objects.get_or_create(user=request.user)
    
    if request.method == 'POST':
        user_form = UserUpdateForm(request.POST, instance=request.user)
        profile_form = ProfileUpdateForm(request.POST, request.FILES, instance=user_profile)
        
        if user_form.is_valid() and profile_form.is_valid():
            user_form.save()
            profile_form.save()
            
            # Log activity
            log_activity(request.user, 'profile_update', "Updated profile information")
            
            messages.success(request, 'Your profile has been updated successfully!')
            return redirect('profile')
        else:
            messages.error(request, 'Please correct the errors below.')
    else:
        user_form = UserUpdateForm(instance=request.user)
        profile_form = ProfileUpdateForm(instance=user_profile)
    
    context = {
        'user_form': user_form,
        'profile_form': profile_form,
        'user_profile': user_profile,
    }
    return render(request, 'profile_edit.html', context)

@login_required
def profile_delete_picture(request):
    """Delete profile picture"""
    if request.method == 'POST':
        user_profile = request.user.userprofile
        if user_profile.profile_picture:
            # Delete the file
            if os.path.isfile(user_profile.profile_picture.path):
                os.remove(user_profile.profile_picture.path)
            user_profile.profile_picture.delete()
            messages.success(request, 'Profile picture removed successfully.')
    return redirect('profile_edit')